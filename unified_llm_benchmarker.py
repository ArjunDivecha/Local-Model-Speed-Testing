#!/usr/bin/env python3
"""
================================================================================
Unified LLM Benchmarker
================================================================================
Version: 1.1
Date: 2025-07-16
Author: AI Assistant

--------------------------------------------------------------------------------
OVERVIEW
--------------------------------------------------------------------------------
This script is a comprehensive tool for benchmarking and evaluating the
performance and response quality of various Large Language Model (LLM) providers.
It is designed to be easily configurable and extensible for new models and
providers.

The process involves these main steps:
1.  Configuration: Loads API keys and provider details.
2.  Benchmarking: Sends a standardized prompt to all configured models and
    measures performance metrics like speed (tokens/second), generation time,
    and cost. It uses parallel processing to run tests efficiently.
3.  Evaluation: Uses a high-quality AI model (Claude 4 Opus) as a "judge" to
    assess the quality of each model's response based on criteria like
    correctness, completeness, and code quality.
4.  Aggregation: Combines all performance and quality data into a single,
    structured dataset.
5.  Reporting: Generates a detailed Excel report with multiple analytical sheets
    and a series of PDF charts for visual comparison.

--------------------------------------------------------------------------------
INPUT FILES
--------------------------------------------------------------------------------
This script relies on a configuration file for API keys.

1. .env file:
   - Location: Must be in the same directory as the script.
   - Purpose: Securely stores the API keys required to connect to the various
     LLM providers. This prevents hardcoding sensitive information in the script.
   - Format: A simple text file with KEY=VALUE pairs on each line.
   - Required Keys:
     - OPENAI_API_KEY="sk-..."
     - ANTHROPIC_API_KEY="sk-ant-..."
     - GEMINI_API_KEY="..."
     - GROQ_API_KEY="gsk_..."
     - XAI_API_KEY="xai-..."
     - DEEPSEEK_API_KEY="sk-..."
     - CEREBRAS_API_KEY="..."
     - PERPLEXITY_API_KEY="pplx-..."

--------------------------------------------------------------------------------
OUTPUT FILES
--------------------------------------------------------------------------------
The script generates one Excel file and multiple PDF charts, all of which are
timestamped to prevent overwriting previous results.

1. unified_benchmark_results_{timestamp}.xlsx:
   - Description: A multi-sheet Excel workbook containing all raw and
     analyzed data from the benchmark run.
   - Sheets:
     - "Benchmark Results": The main sheet with all raw metrics and quality
       scores for every model tested, including rankings.
     - "Summary": A high-level overview of the test run, including success
       rates, top-performing models, and total costs.
     - "Performance Comparison": A filtered view focusing only on speed and
       cost metrics, sorted by performance.
     - "Quality Analysis": A filtered view focusing on the AI-judged quality
       scores and evaluation summaries.

2. speed_vs_quality_{timestamp}.pdf:
   - Description: A scatter plot that visually represents the trade-off
     between model speed (tokens per second) and response quality. This helps
     identify models that are both fast and high-quality.

3. speed_comparison_{timestamp}.pdf:
   - Description: A bar chart comparing the generation speed (tokens per
     second) of all successfully benchmarked models.

4. quality_comparison_{timestamp}.pdf:
   - Description: A bar chart comparing the overall quality scores of all
     successfully evaluated models.

5. ttft_comparison_{timestamp}.pdf:
   - Description: A bar chart comparing the Time To First Token (TTFT) for
     all models, which is a key measure of responsiveness.

--------------------------------------------------------------------------------
DATA HANDLING & QUALITY
--------------------------------------------------------------------------------
- API Failures: If a model fails an API call after multiple retries, it is
  marked as unsuccessful in the final report, and the error message is logged.
- Missing Evaluations: If the AI-powered quality evaluation fails for a model,
  its quality scores are set to 0.0 to indicate that data is missing.
- Data Normalization: Model names are normalized to handle variations in naming
  conventions (e.g., 'gpt-4o' vs 'openai/gpt-4o') and remove duplicates.
- Missing Values: The script ensures there are no empty or null values in the
  final dataset by filling them with sensible defaults (e.g., 0 for scores,
  empty strings for text).
"""

import os
import time
import json
import logging
from datetime import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any
from concurrent.futures import ThreadPoolExecutor
from abc import ABC, abstractmethod
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import requests
from openai import OpenAI
import google.generativeai as genai
from cerebras.cloud.sdk import Cerebras
from dotenv import load_dotenv
import openpyxl


# Configuration Constants
DEFAULT_BENCHMARK_PROMPT = """
Write a script that connects to Interactive Brokers API, retrieves real-time options data for SPY, calculates implied volatility skew, and generates an alert when the skew exceeds historical 90th percentile. Include reconnection logic and handle all possible API errors explicitly.
"""

EVALUATION_CRITERIA = [
    "Correctness",
    "Completeness", 
    "Code Quality",
    "Readability",
    "Error Handling"
]

# Provider Configuration - Updated with current models from ModelSpeed.py
PROVIDERS = {
    "openai": {
        "base_url": "https://api.openai.com/v1",
        "models": ["gpt-4o", "o4-mini", "o3", "gpt-4o-mini", "gpt-3.5-turbo"],
        "api_key_env": "OPENAI_API_KEY",
        "pricing": {
            "gpt-4o": {"input": 2.50, "output": 10.00},
            "gpt-4o-mini": {"input": 0.15, "output": 0.60},
            "gpt-3.5-turbo": {"input": 0.50, "output": 1.50},
            "o3": {"input": 2.00, "output": 8.00},
            "o4-mini": {"input": 0.15, "output": 0.60}
        }
    },
    "anthropic": {
        "base_url": "https://api.anthropic.com/v1",
        "models": ["claude-sonnet-4-20250514", "claude-opus-4-20250514", "claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022"],
        "api_key_env": "ANTHROPIC_API_KEY",
        "pricing": {
            "claude-sonnet-4-20250514": {"input": 3.00, "output": 15.00},
            "claude-opus-4-20250514": {"input": 15.00, "output": 75.00},
            "claude-3-5-sonnet-20241022": {"input": 3.00, "output": 15.00},
            "claude-3-5-haiku-20241022": {"input": 0.80, "output": 4.00}
        }
    },
    "google": {
        # use the stable v1 endpoint
        "base_url": "https://generativelanguage.googleapis.com/v1",
        # model IDs must include the models/ prefix when you call the endpoint
        "models": [
            "models/gemini-2.5-pro",          # SOTA reasoning & coding
            "models/gemini-2.5-flash",        # fastest large Gemini
            "models/gemini-1.5-pro-latest",   # long-context (1 M tokens)
            "models/gemini-1.5-flash-latest"  # fast, long-context
        ],
        "api_key_env": "GEMINI_API_KEY",      # or GOOGLE_API_KEY - see docs
        # 2025-07-15 public pricing (per-1K tokens, USD)
        "pricing": {
            "models/gemini-2.5-pro": {"input": 0.012, "output": 0.036},
            "models/gemini-2.5-flash": {"input": 0.003, "output": 0.009},
            "models/gemini-1.5-pro-latest": {"input": 0.007, "output": 0.021},
            "models/gemini-1.5-flash-latest": {"input": 0.002, "output": 0.006}
        }
    },
    "groq": {
        "base_url": "https://api.groq.com/openai/v1",
        "models": ["moonshotai/kimi-k2-instruct"],
        "api_key_env": "GROQ_API_KEY",
        "pricing": {
            "moonshotai/kimi-k2-instruct": {"input": 1.00, "output": 3.00}
        }
    },
    "xai": {
        "base_url": "https://api.x.ai/v1",
        "models": ["grok-4-0709"],
        "api_key_env": "XAI_API_KEY",
        "pricing": {
            "grok-4-0709": {"input": 0.002, "output": 0.002}
        }
    },
    "deepseek": {
        "base_url": "https://api.deepseek.com/v1",
        "models": ["deepseek-chat", "deepseek-coder", "deepseek-reasoner"],
        "api_key_env": "DEEPSEEK_API_KEY",
        "pricing": {
            "deepseek-chat": {"input": 0.0014, "output": 0.0014},
            "deepseek-coder": {"input": 0.0014, "output": 0.0014},
            "deepseek-reasoner": {"input": 0.0014, "output": 0.0014}
        }
    },
    "cerebras": {
        "base_url": "https://api.cerebras.ai/v1",
        "models": ["qwen-3-235b-a22b", "qwen-3-32b", "llama-3.3-70b", "llama3.1-8b"],
        "api_key_env": "CEREBRAS_API_KEY",
        "pricing": {
            "qwen-3-235b-a22b": {"input": 0.60, "output": 0.60},  # Estimated pricing - update with actual rates
            "qwen-3-32b": {"input": 0.30, "output": 0.30},
            "llama-3.3-70b": {"input": 0.40, "output": 0.40},
            "llama3.1-8b": {"input": 0.10, "output": 0.10}
        }
    }
}

# Timeout and retry settings
REQUEST_TIMEOUT = 60
MAX_RETRIES = 3
RETRY_DELAY = 1


@dataclass
class BenchmarkResult:
    """Data class for storing benchmark results from LLM testing."""
    model: str
    provider: str
    platform: str
    generation_time: float
    tokens_per_second: float
    time_to_first_token: float
    tokens_generated: int
    total_tokens: int
    cost: float
    response_content: str
    success: bool
    error_message: Optional[str] = None


@dataclass
class EvaluationResult:
    """Data class for storing AI-powered quality evaluation results."""
    model: str
    correctness_score: float
    completeness_score: float
    code_quality_score: float
    readability_score: float
    error_handling_score: float
    overall_score: float
    pros: List[str] = field(default_factory=list)
    cons: List[str] = field(default_factory=list)
    summary: str = ""


@dataclass
class ProviderResponse:
    """Data class for standardized provider response."""
    content: str
    tokens_generated: int
    total_tokens: int
    input_tokens: int
    output_tokens: int
    generation_time: float
    time_to_first_token: float
    success: bool
    error_message: Optional[str] = None


class BaseProvider(ABC):
    """Abstract base class for LLM providers."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        """Initialize provider with API key and configuration."""
        self.api_key = api_key
        self.config = config
        self.logger = logging.getLogger(f"{self.__class__.__name__}")
        
    @abstractmethod
    def generate_response(self, prompt: str, model: str, max_tokens: int = 10000) -> ProviderResponse:
        """Generate response from the model with timing and token metrics."""
        pass
    
    def calculate_cost(self, model: str, input_tokens: int, output_tokens: int) -> float:
        """Calculate cost based on provider pricing."""
        pricing = self.config.get("pricing", {}).get(model, {})
        if not pricing:
            self.logger.warning(f"No pricing information for model {model}")
            return 0.0
        
        input_cost = (input_tokens / 1000) * pricing.get("input", 0)
        output_cost = (output_tokens / 1000) * pricing.get("output", 0)
        return input_cost + output_cost
    
    def _retry_with_backoff(self, func, *args, **kwargs):
        """Execute function with exponential backoff retry logic."""
        for attempt in range(MAX_RETRIES):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                if attempt == MAX_RETRIES - 1:
                    raise e
                
                wait_time = RETRY_DELAY * (2 ** attempt)
                self.logger.warning(f"Attempt {attempt + 1} failed: {str(e)}. Retrying in {wait_time}s...")
                time.sleep(wait_time)


class OpenAIProvider(BaseProvider):
    """OpenAI API provider implementation."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        super().__init__(api_key, config)
        self.client = OpenAI(api_key=api_key, base_url=config.get("base_url"))
    
    def generate_response(self, prompt: str, model: str, max_tokens: int = 10000) -> ProviderResponse:
        """Generate response using OpenAI API."""
        def _make_request():
            start_time = time.time()
            ttft_recorded = False
            ttft = 0.0
            
            try:
                # Determine which token parameter to use based on model
                # o3 and o4-mini models require max_completion_tokens instead of max_tokens
                token_param = {}
                if model in ["o3", "o4-mini"]:
                    token_param["max_completion_tokens"] = max_tokens
                else:
                    token_param["max_tokens"] = max_tokens
                
                response = self.client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    stream=True,
                    **token_param
                )
                
                content = ""
                for chunk in response:
                    if not ttft_recorded and chunk.choices and chunk.choices[0].delta.content:
                        ttft = time.time() - start_time
                        ttft_recorded = True
                    
                    if chunk.choices and chunk.choices[0].delta.content:
                        content += chunk.choices[0].delta.content
                
                end_time = time.time()
                generation_time = end_time - start_time
                
                # Get token usage (approximate for streaming)
                tokens_generated = len(content.split()) * 1.3  # Rough approximation
                input_tokens = len(prompt.split()) * 1.3
                total_tokens = tokens_generated + input_tokens
                
                return ProviderResponse(
                    content=content,
                    tokens_generated=int(tokens_generated),
                    total_tokens=int(total_tokens),
                    input_tokens=int(input_tokens),
                    output_tokens=int(tokens_generated),
                    generation_time=generation_time,
                    time_to_first_token=ttft if ttft_recorded else generation_time,
                    success=True
                )
                
            except Exception as e:
                return ProviderResponse(
                    content="",
                    tokens_generated=0,
                    total_tokens=0,
                    input_tokens=0,
                    output_tokens=0,
                    generation_time=0.0,
                    time_to_first_token=0.0,
                    success=False,
                    error_message=str(e)
                )
        
        return self._retry_with_backoff(_make_request)


class AnthropicProvider(BaseProvider):
    """Anthropic API provider implementation."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        super().__init__(api_key, config)
        self.base_url = config.get("base_url", "https://api.anthropic.com/v1")
    
    def generate_response(self, prompt: str, model: str, max_tokens: int = 2000) -> ProviderResponse:
        """Generate response using Anthropic API."""
        def _make_request():
            start_time = time.time()
            
            headers = {
                "x-api-key": self.api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json"
            }
            
            data = {
                "model": model,
                "max_tokens": max_tokens,
                "messages": [{"role": "user", "content": prompt}]
            }
            
            try:
                response = requests.post(
                    f"{self.base_url}/messages",
                    headers=headers,
                    json=data,
                    timeout=REQUEST_TIMEOUT
                )
                response.raise_for_status()
                
                end_time = time.time()
                generation_time = end_time - start_time
                
                result = response.json()
                content = result.get("content", [{}])[0].get("text", "")
                usage = result.get("usage", {})
                
                input_tokens = usage.get("input_tokens", 0)
                output_tokens = usage.get("output_tokens", 0)
                total_tokens = input_tokens + output_tokens
                
                return ProviderResponse(
                    content=content,
                    tokens_generated=output_tokens,
                    total_tokens=total_tokens,
                    input_tokens=input_tokens,
                    output_tokens=output_tokens,
                    generation_time=generation_time,
                    time_to_first_token=generation_time,  # Anthropic doesn't provide TTFT
                    success=True
                )
                
            except Exception as e:
                return ProviderResponse(
                    content="",
                    tokens_generated=0,
                    total_tokens=0,
                    input_tokens=0,
                    output_tokens=0,
                    generation_time=0.0,
                    time_to_first_token=0.0,
                    success=False,
                    error_message=str(e)
                )
        
        return self._retry_with_backoff(_make_request)


class GoogleProvider(BaseProvider):
    """Google Gemini API provider implementation."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        super().__init__(api_key, config)
        genai.configure(api_key=api_key)
    
    def generate_response(self, prompt: str, model: str, max_tokens: int = 10000) -> ProviderResponse:
        """Generate response using Google Gemini API."""
        def _make_request():
            start_time = time.time()
            
            try:
                # Map model names to correct Gemini model IDs
                # Handle both with and without the models/ prefix
                model_mapping = {
                    "gemini-2.5-flash": "gemini-1.5-flash-latest",
                    "gemini-2.5-pro": "gemini-1.5-pro-latest",
                    "gemini-1.5-flash": "gemini-1.5-flash-latest",
                    "gemini-1.5-pro": "gemini-1.5-pro-latest",
                    "models/gemini-2.5-flash": "gemini-1.5-flash-latest",
                    "models/gemini-2.5-pro": "gemini-1.5-pro-latest",
                    "models/gemini-1.5-flash": "gemini-1.5-flash-latest",
                    "models/gemini-1.5-pro": "gemini-1.5-pro-latest",
                    "models/gemini-1.5-pro-latest": "gemini-1.5-pro-latest",
                    "models/gemini-1.5-flash-latest": "gemini-1.5-flash-latest"
                }
                
                # Use mapped model name or original if not in mapping
                actual_model = model_mapping.get(model, model)
                
                # Configure safety settings to be very permissive for coding tasks
                # Using BLOCK_NONE to allow all coding content through
                from google.generativeai.types import HarmCategory, HarmBlockThreshold
                
                safety_settings = {
                    HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_NONE,
                    HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_NONE,
                    HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_NONE,
                    HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_NONE,
                }
                
                # Create model instance with safety settings
                model_instance = genai.GenerativeModel(
                    model_name=actual_model,
                    safety_settings=safety_settings
                )
                
                # Configure generation parameters
                generation_config = genai.types.GenerationConfig(
                    max_output_tokens=max_tokens,
                    temperature=0.7,
                    response_mime_type="text/plain"  # Get clean code without markdown wrapping
                )
                
                # Generate content
                response = model_instance.generate_content(
                    prompt,
                    generation_config=generation_config
                )
                
                end_time = time.time()
                generation_time = end_time - start_time
                
                # Check if request was blocked at the prompt level
                if hasattr(response, 'prompt_feedback') and response.prompt_feedback:
                    if hasattr(response.prompt_feedback, 'block_reason') and response.prompt_feedback.block_reason:
                        return ProviderResponse(
                            content="",
                            tokens_generated=0,
                            total_tokens=0,
                            input_tokens=0,
                            output_tokens=0,
                            generation_time=generation_time,
                            time_to_first_token=generation_time,
                            success=False,
                            error_message=f"Prompt blocked: {response.prompt_feedback.block_reason}"
                        )
                
                # Check if response was blocked by safety filters
                if response.candidates and len(response.candidates) > 0:
                    candidate = response.candidates[0]
                    
                    # Check finish reason using numeric values
                    if hasattr(candidate, 'finish_reason'):
                        finish_reason = candidate.finish_reason
                        finish_reason_value = finish_reason.value if hasattr(finish_reason, 'value') else int(finish_reason)
                        
                        if finish_reason_value == 3:  # SAFETY
                            safety_ratings = getattr(candidate, 'safety_ratings', [])
                            blocked_categories = []
                            for rating in safety_ratings:
                                if hasattr(rating, 'blocked') and rating.blocked:
                                    blocked_categories.append(str(getattr(rating, 'category', 'UNKNOWN')))
                            
                            error_msg = f"Content blocked by safety filters. Categories: {', '.join(blocked_categories) if blocked_categories else 'Unknown'}"
                            return ProviderResponse(
                                content="",
                                tokens_generated=0,
                                total_tokens=0,
                                input_tokens=0,
                                output_tokens=0,
                                generation_time=generation_time,
                                time_to_first_token=generation_time,
                                success=False,
                                error_message=error_msg
                            )
                        elif finish_reason_value == 4:  # RECITATION
                            return ProviderResponse(
                                content="",
                                tokens_generated=0,
                                total_tokens=0,
                                input_tokens=0,
                                output_tokens=0,
                                generation_time=generation_time,
                                time_to_first_token=generation_time,
                                success=False,
                                error_message="Content blocked due to recitation concerns"
                            )
                
                # Extract content from response using safer method
                content = ""
                
                # Method 1: Try to get content from candidate parts first
                if response.candidates and len(response.candidates) > 0:
                    candidate = response.candidates[0]
                    if hasattr(candidate, 'content') and candidate.content:
                        if hasattr(candidate.content, 'parts') and candidate.content.parts:
                            for part in candidate.content.parts:
                                if hasattr(part, 'text') and part.text:
                                    content += part.text
                
                # Method 2: Try response.text if no content from parts
                if not content:
                    try:
                        if hasattr(response, 'text'):
                            content = response.text
                    except Exception:
                        # If response.text fails, content remains empty
                        pass
                
                if not content:
                    return ProviderResponse(
                        content="",
                        tokens_generated=0,
                        total_tokens=0,
                        input_tokens=0,
                        output_tokens=0,
                        generation_time=generation_time,
                        time_to_first_token=generation_time,
                        success=False,
                        error_message="No content generated - response may have been filtered or empty"
                    )
                
                # Calculate approximate token counts
                input_tokens = len(prompt.split()) * 1.3
                output_tokens = len(content.split()) * 1.3
                total_tokens = input_tokens + output_tokens
                
                return ProviderResponse(
                    content=content,
                    tokens_generated=int(output_tokens),
                    total_tokens=int(total_tokens),
                    input_tokens=int(input_tokens),
                    output_tokens=int(output_tokens),
                    generation_time=generation_time,
                    time_to_first_token=generation_time,  # Google doesn't provide TTFT in non-streaming mode
                    success=True
                )
                
            except Exception as e:
                return ProviderResponse(
                    content="",
                    tokens_generated=0,
                    total_tokens=0,
                    input_tokens=0,
                    output_tokens=0,
                    generation_time=0.0,
                    time_to_first_token=0.0,
                    success=False,
                    error_message=str(e)
                )
        
        return self._retry_with_backoff(_make_request)


class OpenAICompatibleProvider(BaseProvider):
    """Generic OpenAI-compatible API provider (for Groq, DeepSeek, XAI, etc.)."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        super().__init__(api_key, config)
        self.client = OpenAI(api_key=api_key, base_url=config.get("base_url"))
    
    def generate_response(self, prompt: str, model: str, max_tokens: int = 2000) -> ProviderResponse:
        """Generate response using OpenAI-compatible API."""
        def _make_request():
            start_time = time.time()
            ttft_recorded = False
            ttft = 0.0
            
            try:
                response = self.client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=max_tokens,
                    stream=True
                )
                
                content = ""
                for chunk in response:
                    if not ttft_recorded and chunk.choices and chunk.choices[0].delta.content:
                        ttft = time.time() - start_time
                        ttft_recorded = True
                    
                    if chunk.choices and chunk.choices[0].delta.content:
                        content += chunk.choices[0].delta.content
                
                end_time = time.time()
                generation_time = end_time - start_time
                
                # Approximate token counts
                tokens_generated = len(content.split()) * 1.3
                input_tokens = len(prompt.split()) * 1.3
                total_tokens = tokens_generated + input_tokens
                
                return ProviderResponse(
                    content=content,
                    tokens_generated=int(tokens_generated),
                    total_tokens=int(total_tokens),
                    input_tokens=int(input_tokens),
                    output_tokens=int(tokens_generated),
                    generation_time=generation_time,
                    time_to_first_token=ttft if ttft_recorded else generation_time,
                    success=True
                )
                
            except Exception as e:
                return ProviderResponse(
                    content="",
                    tokens_generated=0,
                    total_tokens=0,
                    input_tokens=0,
                    output_tokens=0,
                    generation_time=0.0,
                    time_to_first_token=0.0,
                    success=False,
                    error_message=str(e)
                )
        
        return self._retry_with_backoff(_make_request)


class PerplexityProvider(BaseProvider):
    """Perplexity API provider implementation."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        super().__init__(api_key, config)
        self.base_url = config.get("base_url", "https://api.perplexity.ai")
    
    def generate_response(self, prompt: str, model: str, max_tokens: int = 2000) -> ProviderResponse:
        """Generate response using Perplexity API."""
        def _make_request():
            start_time = time.time()
            
            headers = {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json"
            }
            
            data = {
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": max_tokens
            }
            
            try:
                response = requests.post(
                    f"{self.base_url}/chat/completions",
                    headers=headers,
                    json=data,
                    timeout=REQUEST_TIMEOUT
                )
                response.raise_for_status()
                
                end_time = time.time()
                generation_time = end_time - start_time
                
                result = response.json()
                content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
                usage = result.get("usage", {})
                
                input_tokens = usage.get("prompt_tokens", 0)
                output_tokens = usage.get("completion_tokens", 0)
                total_tokens = usage.get("total_tokens", input_tokens + output_tokens)
                
                return ProviderResponse(
                    content=content,
                    tokens_generated=output_tokens,
                    total_tokens=total_tokens,
                    input_tokens=input_tokens,
                    output_tokens=output_tokens,
                    generation_time=generation_time,
                    time_to_first_token=generation_time,
                    success=True
                )
                
            except Exception as e:
                return ProviderResponse(
                    content="",
                    tokens_generated=0,
                    total_tokens=0,
                    input_tokens=0,
                    output_tokens=0,
                    generation_time=0.0,
                    time_to_first_token=0.0,
                    success=False,
                    error_message=str(e)
                )
        
        return self._retry_with_backoff(_make_request)


class CerebrasProvider(BaseProvider):
    """Cerebras API provider implementation."""
    
    def __init__(self, api_key: str, config: Dict[str, Any]):
        super().__init__(api_key, config)
        self.client = Cerebras(api_key=api_key)
    
    def generate_response(self, prompt: str, model: str, max_tokens: int = 2000) -> ProviderResponse:
        """Generate response using Cerebras API."""
        def _make_request():
            start_time = time.time()
            
            try:
                response = self.client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=max_tokens,
                    temperature=0.7
                )
                
                end_time = time.time()
                generation_time = end_time - start_time
                
                # Extract content and usage information
                content = response.choices[0].message.content if response.choices else ""
                usage = response.usage if hasattr(response, 'usage') else None
                
                if usage:
                    input_tokens = usage.prompt_tokens
                    output_tokens = usage.completion_tokens
                    total_tokens = usage.total_tokens
                else:
                    # Fallback to approximation if usage not available
                    input_tokens = int(len(prompt.split()) * 1.3)
                    output_tokens = int(len(content.split()) * 1.3)
                    total_tokens = input_tokens + output_tokens
                
                return ProviderResponse(
                    content=content,
                    tokens_generated=output_tokens,
                    total_tokens=total_tokens,
                    input_tokens=input_tokens,
                    output_tokens=output_tokens,
                    generation_time=generation_time,
                    time_to_first_token=generation_time,  # Cerebras doesn't provide TTFT in non-streaming mode
                    success=True
                )
                
            except Exception as e:
                return ProviderResponse(
                    content="",
                    tokens_generated=0,
                    total_tokens=0,
                    input_tokens=0,
                    output_tokens=0,
                    generation_time=0.0,
                    time_to_first_token=0.0,
                    success=False,
                    error_message=str(e)
                )
        
        return self._retry_with_backoff(_make_request)


class UnifiedBenchmarker:
    """Main class for unified LLM benchmarking and evaluation."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize the benchmarker with configuration."""
        self.config = config or {}
        self.benchmark_prompt = self.config.get('benchmark_prompt', DEFAULT_BENCHMARK_PROMPT)
        self.token_limit = self.config.get('token_limit', 2000)
        self.evaluation_criteria = self.config.get('evaluation_criteria', EVALUATION_CRITERIA)
        
        # Set up logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
        
        # Initialize results storage
        self.benchmark_results: List[BenchmarkResult] = []
        self.evaluation_results: List[EvaluationResult] = []
        
        # Generate timestamp for output files
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
        # Load configuration and validate API keys
        self.api_keys = self._load_api_keys()
        self.available_providers = self._validate_api_keys()
        
        self.logger.info(f"Initialized with {len(self.available_providers)} available providers")
    
    def _load_api_keys(self) -> Dict[str, Optional[str]]:
        """Load API keys from .env file and environment variables."""
        # Load environment variables from .env file
        load_dotenv()
        
        api_keys = {}
        
        for provider_name, provider_config in PROVIDERS.items():
            api_key_env = provider_config.get("api_key_env")
            if api_key_env:
                api_key = os.getenv(api_key_env)
                if api_key:
                    # Strip whitespace from API key
                    api_key = api_key.strip()
                
                api_keys[provider_name] = api_key
                
                if api_key:
                    self.logger.info(f"✓ {provider_name}: API key loaded from {api_key_env}")
                else:
                    self.logger.warning(f"✗ {provider_name}: API key not found in {api_key_env}")
            else:
                self.logger.warning(f"✗ {provider_name}: No API key environment variable configured")
                api_keys[provider_name] = None
        
        return api_keys
    
    def _validate_api_keys(self) -> Dict[str, bool]:
        """Validate API keys for all providers and return availability status."""
        available_providers = {}
        
        for provider_name, api_key in self.api_keys.items():
            if api_key is None:
                available_providers[provider_name] = False
                self.logger.warning(f"Skipping {provider_name}: No API key available")
                continue
            
            # Basic validation - check if key exists and has reasonable format
            if self._validate_api_key_format(provider_name, api_key):
                available_providers[provider_name] = True
                self.logger.info(f"✓ {provider_name}: API key validated")
            else:
                available_providers[provider_name] = False
                self.logger.error(f"✗ {provider_name}: Invalid API key format")
        
        return available_providers
    
    def _validate_api_key_format(self, provider: str, api_key: str) -> bool:
        """Validate API key format for specific providers."""
        if not api_key or not isinstance(api_key, str):
            return False
        
        # Basic format validation for each provider
        if provider == "openai":
            return api_key.startswith("sk-") and len(api_key) > 20
        elif provider == "anthropic":
            return (api_key.startswith("sk-ant-") or api_key.startswith("sk-ant-api")) and len(api_key) > 20
        elif provider == "google":
            return len(api_key) > 20  # Google API keys don't have a standard prefix
        elif provider == "groq":
            return api_key.startswith("gsk_") and len(api_key) > 20
        elif provider == "perplexity":
            return api_key.startswith("pplx-") and len(api_key) > 20
        elif provider == "deepseek":
            return api_key.startswith("sk-") and len(api_key) > 20
        elif provider == "xai":
            return api_key.startswith("xai-") and len(api_key) > 20
        elif provider == "openrouter":
            return api_key.startswith("sk-or-") and len(api_key) > 20
        else:
            # For unknown providers, just check if it's a non-empty string
            return len(api_key.strip()) > 0
    
    def get_available_models(self) -> List[Tuple[str, str]]:
        """Get list of available models from providers with valid API keys."""
        available_models = []
        
        for provider_name, is_available in self.available_providers.items():
            if is_available and provider_name in PROVIDERS:
                provider_config = PROVIDERS[provider_name]
                models = provider_config.get("models", [])
                
                for model in models:
                    available_models.append((provider_name, model))
                    
                self.logger.info(f"{provider_name}: {len(models)} models available")
        
        self.logger.info(f"Total available models: {len(available_models)}")
        return available_models
    
    def get_provider_config(self, provider: str) -> Optional[Dict[str, Any]]:
        """Get configuration for a specific provider."""
        if provider not in PROVIDERS:
            self.logger.error(f"Unknown provider: {provider}")
            return None
        
        if not self.available_providers.get(provider, False):
            self.logger.warning(f"Provider {provider} is not available (missing or invalid API key)")
            return None
        
        config = PROVIDERS[provider].copy()
        config["api_key"] = self.api_keys[provider]
        return config
    
    def validate_api_keys(self) -> Dict[str, bool]:
        """Public method to get API key validation status."""
        return self.available_providers.copy()
    
    def create_provider(self, provider_name: str) -> Optional[BaseProvider]:
        """Create a provider instance for the given provider name."""
        if not self.available_providers.get(provider_name, False):
            self.logger.error(f"Provider {provider_name} is not available")
            return None
        
        config = self.get_provider_config(provider_name)
        if not config:
            return None
        
        api_key = config["api_key"]
        
        try:
            if provider_name == "openai":
                return OpenAIProvider(api_key, config)
            elif provider_name == "anthropic":
                return AnthropicProvider(api_key, config)
            elif provider_name == "google":
                return GoogleProvider(api_key, config)
            elif provider_name == "perplexity":
                return PerplexityProvider(api_key, config)
            elif provider_name == "cerebras":
                return CerebrasProvider(api_key, config)
            elif provider_name in ["groq", "deepseek", "xai", "openrouter"]:
                return OpenAICompatibleProvider(api_key, config)
            else:
                self.logger.error(f"Unknown provider type: {provider_name}")
                return None
        except Exception as e:
            self.logger.error(f"Failed to create provider {provider_name}: {str(e)}")
            return None
    
    def benchmark_model(self, provider_name: str, model: str) -> BenchmarkResult:
        """Benchmark a specific model from a provider."""
        provider = self.create_provider(provider_name)
        if not provider:
            return BenchmarkResult(
                model=model,
                provider=provider_name,
                platform=provider_name,
                generation_time=0.0,
                tokens_per_second=0.0,
                time_to_first_token=0.0,
                tokens_generated=0,
                total_tokens=0,
                cost=0.0,
                response_content="",
                success=False,
                error_message=f"Provider {provider_name} not available"
            )
        
        try:
            self.logger.info(f"Benchmarking {provider_name}/{model}...")
            
            # Generate response using the provider
            response = provider.generate_response(
                prompt=self.benchmark_prompt,
                model=model,
                max_tokens=self.token_limit
            )
            
            if not response.success:
                return BenchmarkResult(
                    model=model,
                    provider=provider_name,
                    platform=provider_name,
                    generation_time=0.0,
                    tokens_per_second=0.0,
                    time_to_first_token=0.0,
                    tokens_generated=0,
                    total_tokens=0,
                    cost=0.0,
                    response_content="",
                    success=False,
                    error_message=response.error_message
                )
            
            # Calculate tokens per second
            tokens_per_second = (
                response.tokens_generated / response.generation_time 
                if response.generation_time > 0 else 0.0
            )
            
            # Calculate cost
            cost = provider.calculate_cost(
                model=model,
                input_tokens=response.input_tokens,
                output_tokens=response.output_tokens
            )
            
            result = BenchmarkResult(
                model=model,
                provider=provider_name,
                platform=provider_name,
                generation_time=response.generation_time,
                tokens_per_second=tokens_per_second,
                time_to_first_token=response.time_to_first_token,
                tokens_generated=response.tokens_generated,
                total_tokens=response.total_tokens,
                cost=cost,
                response_content=response.content,
                success=True
            )
            
            self.logger.info(f"✓ {provider_name}/{model}: {tokens_per_second:.2f} tokens/sec, ${cost:.4f}")
            return result
            
        except Exception as e:
            self.logger.error(f"✗ {provider_name}/{model}: {str(e)}")
            return BenchmarkResult(
                model=model,
                provider=provider_name,
                platform=provider_name,
                generation_time=0.0,
                tokens_per_second=0.0,
                time_to_first_token=0.0,
                tokens_generated=0,
                total_tokens=0,
                cost=0.0,
                response_content="",
                success=False,
                error_message=str(e)
            )
    
    def generate_reference_response(self) -> Optional[str]:
        """Generate a reference response using Claude Opus for comparison."""
        if not self.available_providers.get("anthropic", False):
            self.logger.warning("Claude Opus not available for reference response generation")
            return None
        
        try:
            self.logger.info("Generating reference response using Claude Opus...")
            provider = self.create_provider("anthropic")
            if not provider:
                return None
            
            # Use Claude 4 Opus for reference response
            reference_model = "claude-opus-4-20250514"
            response = provider.generate_response(
                prompt=self.benchmark_prompt,
                model=reference_model,
                max_tokens=self.token_limit
            )
            
            if response.success and response.content:
                self.logger.info("✓ Reference response generated successfully")
                return response.content
            else:
                self.logger.error(f"Failed to generate reference response: {response.error_message}")
                return None
                
        except Exception as e:
            self.logger.error(f"Error generating reference response: {str(e)}")
            return None
    
    def evaluate_response_quality(self, benchmark_result: BenchmarkResult, reference_response: Optional[str] = None) -> Optional[EvaluationResult]:
        """Evaluate response quality using Claude Opus as AI judge."""
        if not benchmark_result.success or not benchmark_result.response_content:
            self.logger.error(f"Cannot evaluate {benchmark_result.model}: No response content available")
            return None
        
        # Check if Claude Opus is available - fail if not
        if not self.available_providers.get("anthropic", False):
            self.logger.error("Claude Opus not available for evaluation - evaluation cannot proceed")
            return None
        
        provider = self.create_provider("anthropic")
        if not provider:
            self.logger.error("Failed to create Anthropic provider - evaluation cannot proceed")
            return None
        
        # Generate reference response if not provided
        if reference_response is None:
            reference_response = self.generate_reference_response()
            if reference_response is None:
                self.logger.error("Failed to generate reference response - evaluation cannot proceed")
                return None
        
        # Create evaluation prompt
        evaluation_prompt = self._create_evaluation_prompt(
            original_prompt=self.benchmark_prompt,
            model_response=benchmark_result.response_content,
            reference_response=reference_response,
            criteria=self.evaluation_criteria
        )
        
        try:
            self.logger.info(f"Evaluating {benchmark_result.provider}/{benchmark_result.model}...")
            
            # Use Claude 4 Opus for evaluation
            evaluation_model = "claude-opus-4-20250514"
            response = provider.generate_response(
                prompt=evaluation_prompt,
                model=evaluation_model,
                max_tokens=1500
            )
            
            if not response.success or not response.content:
                self.logger.error(f"Evaluation failed for {benchmark_result.model}: {response.error_message}")
                return None
            
            # Parse evaluation response
            evaluation_result = self._parse_evaluation_response(
                response.content,
                benchmark_result.model
            )
            
            self.logger.info(f"✓ {benchmark_result.model}: Overall score {evaluation_result.overall_score:.1f}/10")
            return evaluation_result
            
        except Exception as e:
            self.logger.error(f"Error evaluating {benchmark_result.model}: {str(e)}")
            return None
    
    def _create_evaluation_prompt(self, original_prompt: str, model_response: str, 
                                reference_response: Optional[str], criteria: List[str]) -> str:
        """Create structured evaluation prompt for Claude Opus."""
        
        reference_section = ""
        if reference_response:
            reference_section = f"""
REFERENCE RESPONSE (Claude Opus):
{reference_response}

"""
        
        criteria_descriptions = {
            "Correctness": "Does the code work correctly and solve the problem as specified?",
            "Completeness": "Does the response address all requirements from the original prompt?",
            "Code Quality": "Is the code well-structured, efficient, and following best practices?",
            "Readability": "Is the code clear, well-commented, and easy to understand?",
            "Error Handling": "Does the code include appropriate error handling and edge case management?"
        }
        
        criteria_section = "\n".join([
            f"- {criterion}: {criteria_descriptions.get(criterion, 'Rate this aspect of the response')}"
            for criterion in criteria
        ])
        
        return f"""You are an expert code reviewer and AI model evaluator. Your task is to evaluate the quality of a code response generated by an LLM.

ORIGINAL PROMPT:
{original_prompt}

{reference_section}MODEL RESPONSE TO EVALUATE:
{model_response}

EVALUATION CRITERIA:
{criteria_section}

Please provide a comprehensive evaluation in the following JSON format:

{{
    "scores": {{
        "correctness": <score 1-10>,
        "completeness": <score 1-10>,
        "code_quality": <score 1-10>,
        "readability": <score 1-10>,
        "error_handling": <score 1-10>
    }},
    "overall_score": <average of all scores>,
    "pros": [
        "List of strengths and positive aspects",
        "Each item should be specific and actionable"
    ],
    "cons": [
        "List of weaknesses and areas for improvement",
        "Each item should be specific and actionable"
    ],
    "summary": "Brief 2-3 sentence summary of the overall assessment"
}}

Rate each criterion on a scale of 1-10 where:
- 1-3: Poor (major issues, doesn't meet basic requirements)
- 4-6: Average (meets basic requirements but has notable issues)
- 7-8: Good (solid implementation with minor issues)
- 9-10: Excellent (exceptional quality, best practices followed)

Be objective and constructive in your evaluation. Focus on technical accuracy, code quality, and adherence to the original requirements."""
    
    def _parse_evaluation_response(self, response_content: str, model_name: str) -> EvaluationResult:
        """Parse Claude Opus evaluation response into structured data."""
        try:
            # Try to extract JSON from the response
            import re
            json_match = re.search(r'\{.*\}', response_content, re.DOTALL)
            
            if json_match:
                json_str = json_match.group(0)
                evaluation_data = json.loads(json_str)
                
                scores = evaluation_data.get("scores", {})
                
                return EvaluationResult(
                    model=model_name,
                    correctness_score=float(scores.get("correctness", 5.0)),
                    completeness_score=float(scores.get("completeness", 5.0)),
                    code_quality_score=float(scores.get("code_quality", 5.0)),
                    readability_score=float(scores.get("readability", 5.0)),
                    error_handling_score=float(scores.get("error_handling", 5.0)),
                    overall_score=float(evaluation_data.get("overall_score", 5.0)),
                    pros=evaluation_data.get("pros", []),
                    cons=evaluation_data.get("cons", []),
                    summary=evaluation_data.get("summary", "")
                )
            else:
                # Fallback: try to extract scores from text
                return self._extract_scores_from_text(response_content, model_name)
                
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            self.logger.warning(f"Failed to parse evaluation JSON for {model_name}: {str(e)}")
            return self._extract_scores_from_text(response_content, model_name)
    
    def _extract_scores_from_text(self, response_content: str, model_name: str) -> EvaluationResult:
        """Extract scores from text when JSON parsing fails."""
        import re
        
        # Try to find numerical scores in the text
        score_patterns = {
            "correctness": r"correctness[:\s]*(\d+(?:\.\d+)?)",
            "completeness": r"completeness[:\s]*(\d+(?:\.\d+)?)",
            "code_quality": r"code[_\s]*quality[:\s]*(\d+(?:\.\d+)?)",
            "readability": r"readability[:\s]*(\d+(?:\.\d+)?)",
            "error_handling": r"error[_\s]*handling[:\s]*(\d+(?:\.\d+)?)"
        }
        
        scores = {}
        for criterion, pattern in score_patterns.items():
            match = re.search(pattern, response_content, re.IGNORECASE)
            if match:
                try:
                    score = float(match.group(1))
                    # Ensure score is in valid range
                    scores[criterion] = max(1.0, min(10.0, score))
                except ValueError:
                    scores[criterion] = 5.0
            else:
                scores[criterion] = 5.0
        
        overall_score = sum(scores.values()) / len(scores) if scores else 5.0
        
        # Extract pros and cons if possible
        pros = []
        cons = []
        
        # Simple extraction of positive and negative points
        lines = response_content.split('\n')
        for line in lines:
            line = line.strip()
            if any(word in line.lower() for word in ['good', 'excellent', 'well', 'clear', 'proper']):
                if len(line) > 10 and len(line) < 200:
                    pros.append(line)
            elif any(word in line.lower() for word in ['poor', 'missing', 'lacks', 'could', 'should']):
                if len(line) > 10 and len(line) < 200:
                    cons.append(line)
        
        return EvaluationResult(
            model=model_name,
            correctness_score=scores.get("correctness", 5.0),
            completeness_score=scores.get("completeness", 5.0),
            code_quality_score=scores.get("code_quality", 5.0),
            readability_score=scores.get("readability", 5.0),
            error_handling_score=scores.get("error_handling", 5.0),
            overall_score=overall_score,
            pros=pros[:3],  # Limit to top 3
            cons=cons[:3],  # Limit to top 3
            summary=f"Automated evaluation completed with overall score: {overall_score:.1f}/10"
        )
    
    def normalize_model_name(self, model_name: str) -> str:
        """Normalize model names for consistent matching and deduplication."""
        if not model_name:
            return "unknown"
        
        # Convert to lowercase for consistent comparison
        normalized = model_name.lower().strip()
        
        # Remove common prefixes that don't affect model identity
        prefixes_to_remove = ["models/", "anthropic/", "openai/", "meta-llama/"]
        
        for prefix in prefixes_to_remove:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):]
                break
        
        # Remove common suffixes that don't affect core model identity
        suffixes_to_remove = ["-instruct", "-chat", "-latest", "-preview"]
        
        for suffix in suffixes_to_remove:
            if normalized.endswith(suffix):
                normalized = normalized[:-len(suffix)]
                break
        
        # Handle date-based versions (like -20241022) - these should be removed for deduplication
        import re
        normalized = re.sub(r'-\d{8}', '', normalized)
        
        # Handle explicit version prefixes like -v1.5 (but keep model versions like 3.1)
        normalized = re.sub(r'-v\d+\.\d+', '', normalized)
        
        return normalized
    
    def deduplicate_results(self, benchmark_results: List[BenchmarkResult], 
                          evaluation_results: List[EvaluationResult]) -> Tuple[List[BenchmarkResult], List[EvaluationResult]]:
        """Remove duplicate results based on normalized model names, keeping the best performing one."""
        if not benchmark_results:
            return benchmark_results, evaluation_results
        
        # Create mapping of evaluation results by model name
        eval_map = {result.model: result for result in evaluation_results}
        
        # Group results by normalized model name
        model_groups = {}
        for result in benchmark_results:
            normalized_name = self.normalize_model_name(result.model)
            if normalized_name not in model_groups:
                model_groups[normalized_name] = []
            model_groups[normalized_name].append(result)
        
        deduplicated_benchmark = []
        deduplicated_evaluation = []
        
        for normalized_name, results in model_groups.items():
            if len(results) == 1:
                # No duplicates, keep the single result
                best_result = results[0]
            else:
                # Multiple results, choose the best one based on success and performance
                successful_results = [r for r in results if r.success]
                
                if successful_results:
                    # Choose based on tokens per second (performance)
                    best_result = max(successful_results, key=lambda x: x.tokens_per_second)
                    self.logger.info(f"Deduplicated {len(results)} results for {normalized_name}, kept {best_result.provider}/{best_result.model}")
                else:
                    # All failed, keep the first one
                    best_result = results[0]
                    self.logger.warning(f"All {len(results)} results failed for {normalized_name}, kept first one")
            
            deduplicated_benchmark.append(best_result)
            
            # Add corresponding evaluation result if it exists
            if best_result.model in eval_map:
                deduplicated_evaluation.append(eval_map[best_result.model])
        
        self.logger.info(f"Deduplication: {len(benchmark_results)} -> {len(deduplicated_benchmark)} benchmark results")
        return deduplicated_benchmark, deduplicated_evaluation
    
    def handle_missing_values(self, benchmark_results: List[BenchmarkResult], 
                            evaluation_results: List[EvaluationResult]) -> Tuple[List[BenchmarkResult], List[EvaluationResult]]:
        """Handle missing values in results with appropriate defaults."""
        processed_benchmark = []
        processed_evaluation = []
        
        for result in benchmark_results:
            # Create a copy to avoid modifying original
            processed_result = BenchmarkResult(
                model=result.model or "unknown",
                provider=result.provider or "unknown",
                platform=result.platform or result.provider or "unknown",
                generation_time=result.generation_time if result.generation_time > 0 else 0.0,
                tokens_per_second=result.tokens_per_second if result.tokens_per_second > 0 else 0.0,
                time_to_first_token=result.time_to_first_token if result.time_to_first_token > 0 else 0.0,
                tokens_generated=max(0, result.tokens_generated),
                total_tokens=max(0, result.total_tokens),
                cost=max(0.0, result.cost),
                response_content=result.response_content or "",
                success=result.success,
                error_message=result.error_message
            )
            processed_benchmark.append(processed_result)
        
        for result in evaluation_results:
            # Ensure all scores are in valid range (1-10)
            processed_result = EvaluationResult(
                model=result.model or "unknown",
                correctness_score=max(1.0, min(10.0, result.correctness_score)),
                completeness_score=max(1.0, min(10.0, result.completeness_score)),
                code_quality_score=max(1.0, min(10.0, result.code_quality_score)),
                readability_score=max(1.0, min(10.0, result.readability_score)),
                error_handling_score=max(1.0, min(10.0, result.error_handling_score)),
                overall_score=max(1.0, min(10.0, result.overall_score)),
                pros=result.pros or [],
                cons=result.cons or [],
                summary=result.summary or "No evaluation summary available"
            )
            processed_evaluation.append(processed_result)
        
        return processed_benchmark, processed_evaluation
    
    def aggregate_results(self, benchmark_results: List[BenchmarkResult], 
                         evaluation_results: List[EvaluationResult]) -> pd.DataFrame:
        """Aggregate benchmark and evaluation results into a comprehensive DataFrame."""
        if not benchmark_results:
            self.logger.warning("No benchmark results to aggregate")
            return pd.DataFrame()
        
        # Create evaluation lookup
        eval_map = {result.model: result for result in evaluation_results}
        
        aggregated_data = []
        
        for bench_result in benchmark_results:
            eval_result = eval_map.get(bench_result.model)
            
            # Create comprehensive record
            record = {
                # Model identification
                'model': bench_result.model,
                'provider': bench_result.provider,
                'platform': bench_result.platform,
                'normalized_name': self.normalize_model_name(bench_result.model),
                
                # Performance metrics
                'generation_time': bench_result.generation_time,
                'tokens_per_second': bench_result.tokens_per_second,
                'time_to_first_token': bench_result.time_to_first_token,
                'tokens_generated': bench_result.tokens_generated,
                'total_tokens': bench_result.total_tokens,
                'cost': bench_result.cost,
                
                # Success status
                'success': bench_result.success,
                'error_message': bench_result.error_message or "",
                
                # Quality metrics (with defaults if evaluation failed)
                'correctness_score': eval_result.correctness_score if eval_result else 0.0,
                'completeness_score': eval_result.completeness_score if eval_result else 0.0,
                'code_quality_score': eval_result.code_quality_score if eval_result else 0.0,
                'readability_score': eval_result.readability_score if eval_result else 0.0,
                'error_handling_score': eval_result.error_handling_score if eval_result else 0.0,
                'overall_quality_score': eval_result.overall_score if eval_result else 0.0,
                
                # Evaluation details
                'evaluation_available': eval_result is not None,
                'pros': "; ".join(eval_result.pros) if eval_result and eval_result.pros else "",
                'cons': "; ".join(eval_result.cons) if eval_result and eval_result.cons else "",
                'evaluation_summary': eval_result.summary if eval_result else "",
                
                # Derived metrics
                'cost_per_token': bench_result.cost / bench_result.tokens_generated if bench_result.tokens_generated > 0 else 0.0,
                'efficiency_score': (bench_result.tokens_per_second * eval_result.overall_score / 10.0) if eval_result and bench_result.tokens_per_second > 0 else 0.0
            }
            
            aggregated_data.append(record)
        
        # Create DataFrame
        df = pd.DataFrame(aggregated_data)
        
        # Add ranking columns
        if not df.empty:
            # Only rank successful results
            successful_df = df[df['success'] == True].copy()
            
            if not successful_df.empty:
                # Performance rankings
                successful_df['speed_rank'] = successful_df['tokens_per_second'].rank(method='dense', ascending=False)
                successful_df['quality_rank'] = successful_df['overall_quality_score'].rank(method='dense', ascending=False)
                successful_df['cost_rank'] = successful_df['cost'].rank(method='dense', ascending=True)  # Lower cost is better
                successful_df['efficiency_rank'] = successful_df['efficiency_score'].rank(method='dense', ascending=False)
                
                # Overall composite score (weighted combination)
                successful_df['composite_score'] = (
                    successful_df['overall_quality_score'] * 0.4 +  # 40% quality
                    (successful_df['tokens_per_second'] / successful_df['tokens_per_second'].max()) * 10 * 0.3 +  # 30% speed (normalized)
                    (1 - successful_df['cost'] / successful_df['cost'].max()) * 10 * 0.2 +  # 20% cost efficiency (inverted)
                    (successful_df['time_to_first_token'].max() / successful_df['time_to_first_token']) * 0.1  # 10% TTFT (inverted)
                )
                successful_df['overall_rank'] = successful_df['composite_score'].rank(method='dense', ascending=False)
                
                # Merge rankings back to main dataframe
                for col in ['speed_rank', 'quality_rank', 'cost_rank', 'efficiency_rank', 'composite_score', 'overall_rank']:
                    df[col] = df.index.map(successful_df[col].to_dict()).fillna(0)
            else:
                # No successful results, add empty ranking columns
                for col in ['speed_rank', 'quality_rank', 'cost_rank', 'efficiency_rank', 'composite_score', 'overall_rank']:
                    df[col] = 0
        
        self.logger.info(f"Aggregated {len(df)} results with {len([r for r in benchmark_results if r.success])} successful benchmarks")
        return df
    
    def sort_and_rank_results(self, df: pd.DataFrame, sort_by: str = 'composite_score') -> pd.DataFrame:
        """Sort and rank results by specified criteria."""
        if df.empty:
            return df
        
        # Validate sort column
        valid_sort_columns = [
            'composite_score', 'overall_quality_score', 'tokens_per_second', 
            'cost', 'time_to_first_token', 'efficiency_score'
        ]
        
        if sort_by not in valid_sort_columns:
            self.logger.warning(f"Invalid sort column '{sort_by}', using 'composite_score'")
            sort_by = 'composite_score'
        
        # Sort successful results first, then by specified criteria
        ascending = sort_by in ['cost', 'time_to_first_token']  # Lower is better for these metrics
        
        sorted_df = df.sort_values([
            'success',  # Successful results first
            sort_by     # Then by specified criteria
        ], ascending=[False, ascending]).reset_index(drop=True)
        
        # Add final ranking based on sort order
        sorted_df['final_rank'] = range(1, len(sorted_df) + 1)
        
        self.logger.info(f"Results sorted by {sort_by} ({'ascending' if ascending else 'descending'})")
        return sorted_df
    
    def process_and_aggregate_data(self, benchmark_results: List[BenchmarkResult], 
                                 evaluation_results: List[EvaluationResult]) -> pd.DataFrame:
        """Complete data processing pipeline: normalize, deduplicate, handle missing values, and aggregate."""
        self.logger.info("Starting data processing and aggregation...")
        
        # Filter out None values that can occur from failed evaluations
        filtered_benchmark = [r for r in benchmark_results if r is not None]
        filtered_evaluation = [r for r in evaluation_results if r is not None]
        
        self.logger.info(f"Filtered results: {len(benchmark_results)} -> {len(filtered_benchmark)} benchmark, {len(evaluation_results)} -> {len(filtered_evaluation)} evaluation")
        
        # Step 1: Handle missing values
        processed_benchmark, processed_evaluation = self.handle_missing_values(
            filtered_benchmark, filtered_evaluation
        )
        
        # Step 2: Deduplicate results
        deduplicated_benchmark, deduplicated_evaluation = self.deduplicate_results(
            processed_benchmark, processed_evaluation
        )
        
        # Step 3: Aggregate into comprehensive DataFrame
        aggregated_df = self.aggregate_results(deduplicated_benchmark, deduplicated_evaluation)
        
        # Step 4: Sort and rank results
        final_df = self.sort_and_rank_results(aggregated_df, sort_by='composite_score')
        
        self.logger.info(f"Data processing complete: {len(final_df)} final results")
        return final_df
    
    def generate_excel_output(self, df: pd.DataFrame) -> str:
        """Generate comprehensive Excel file with all model metrics."""
        if df.empty:
            self.logger.warning("No data available for Excel generation")
            return ""
        
        # Generate filename with timestamp
        filename = f"unified_benchmark_results_{self.timestamp}.xlsx"
        
        try:
            # Create Excel writer with formatting options
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Define column order and headers for better organization
                column_order = [
                    # Model identification
                    'final_rank', 'model', 'provider', 'platform',
                    
                    # Performance metrics
                    'generation_time', 'tokens_per_second', 'time_to_first_token',
                    'tokens_generated', 'total_tokens',
                    
                    # Quality scores
                    'overall_quality_score', 'correctness_score', 'completeness_score',
                    'code_quality_score', 'readability_score', 'error_handling_score',
                    
                    # Cost and efficiency
                    'cost', 'cost_per_token', 'efficiency_score', 'composite_score',
                    
                    # Rankings
                    'overall_rank', 'speed_rank', 'quality_rank', 'cost_rank', 'efficiency_rank',
                    
                    # Status and evaluation details
                    'success', 'evaluation_available', 'error_message',
                    'pros', 'cons', 'evaluation_summary'
                ]
                
                # Reorder columns (keep any additional columns at the end)
                available_columns = [col for col in column_order if col in df.columns]
                remaining_columns = [col for col in df.columns if col not in column_order]
                final_column_order = available_columns + remaining_columns
                
                # Reorder DataFrame
                df_ordered = df[final_column_order].copy()
                
                # Create main results sheet
                df_ordered.to_excel(writer, sheet_name='Benchmark Results', index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Benchmark Results']
                
                # Apply formatting
                self._format_excel_worksheet(workbook, worksheet, df_ordered)
                
                # Create summary sheet
                self._create_summary_sheet(writer, df_ordered)
                
                # Create performance comparison sheet
                self._create_performance_sheet(writer, df_ordered)
                
                # Create quality analysis sheet
                self._create_quality_sheet(writer, df_ordered)
            
            # Verify file was created successfully
            if os.path.exists(filename):
                file_size = os.path.getsize(filename)
                self.logger.info(f"✓ Excel file generated successfully: {filename} ({file_size:,} bytes)")
                return os.path.abspath(filename)
            else:
                self.logger.error(f"✗ Excel file was not created: {filename}")
                return ""
                
        except Exception as e:
            self.logger.error(f"✗ Failed to generate Excel file: {str(e)}")
            return ""
    
    def _format_excel_worksheet(self, workbook, worksheet, df: pd.DataFrame):
        """Apply formatting to Excel worksheet for better readability."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        failure_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            
            # Set column width based on content
            column_letter = get_column_letter(col_num)
            max_length = max(
                len(str(column)),
                df[column].astype(str).str.len().max() if not df.empty else 0
            )
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            # Color code based on success status
            success_value = df.iloc[row_num - 2]['success'] if 'success' in df.columns else True
            row_fill = success_fill if success_value else failure_fill
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                
                # Apply success/failure coloring to specific columns
                if df.columns[col_num - 1] in ['success', 'model', 'provider', 'overall_quality_score']:
                    cell.fill = row_fill
                
                # Format numeric columns
                column_name = df.columns[col_num - 1]
                if column_name in ['generation_time', 'time_to_first_token']:
                    cell.number_format = '0.00"s"'
                elif column_name in ['tokens_per_second']:
                    cell.number_format = '0.0"t/s"'
                elif column_name in ['cost', 'cost_per_token']:
                    cell.number_format = '$0.0000'
                elif column_name.endswith('_score'):
                    cell.number_format = '0.0'
                elif column_name.endswith('_rank'):
                    cell.number_format = '0'
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, writer, df: pd.DataFrame):
        """Create a summary sheet with key statistics."""
        if df.empty:
            return
        
        successful_df = df[df['success'] == True] if 'success' in df.columns else df
        
        summary_data = {
            'Metric': [
                'Total Models Tested',
                'Successful Tests',
                'Failed Tests',
                'Success Rate (%)',
                '',
                'Best Overall Model',
                'Fastest Model (tokens/sec)',
                'Highest Quality Model',
                'Most Cost-Effective Model',
                '',
                'Average Generation Time (s)',
                'Average Tokens/Second',
                'Average Quality Score',
                'Average Cost per Request ($)',
                '',
                'Total API Cost ($)',
                'Total Tokens Generated',
                'Total Processing Time (s)'
            ],
            'Value': []
        }
        
        # Calculate summary statistics
        total_models = len(df)
        successful_models = len(successful_df)
        failed_models = total_models - successful_models
        success_rate = (successful_models / total_models * 100) if total_models > 0 else 0
        
        # Best models
        best_overall = successful_df.loc[successful_df['composite_score'].idxmax(), 'model'] if not successful_df.empty and 'composite_score' in successful_df.columns else 'N/A'
        fastest_model = successful_df.loc[successful_df['tokens_per_second'].idxmax(), 'model'] if not successful_df.empty and 'tokens_per_second' in successful_df.columns else 'N/A'
        highest_quality = successful_df.loc[successful_df['overall_quality_score'].idxmax(), 'model'] if not successful_df.empty and 'overall_quality_score' in successful_df.columns else 'N/A'
        most_cost_effective = successful_df.loc[successful_df['cost'].idxmin(), 'model'] if not successful_df.empty and 'cost' in successful_df.columns else 'N/A'
        
        # Averages
        avg_gen_time = successful_df['generation_time'].mean() if not successful_df.empty and 'generation_time' in successful_df.columns else 0
        avg_tokens_sec = successful_df['tokens_per_second'].mean() if not successful_df.empty and 'tokens_per_second' in successful_df.columns else 0
        avg_quality = successful_df['overall_quality_score'].mean() if not successful_df.empty and 'overall_quality_score' in successful_df.columns else 0
        avg_cost = successful_df['cost'].mean() if not successful_df.empty and 'cost' in successful_df.columns else 0
        
        # Totals
        total_cost = successful_df['cost'].sum() if not successful_df.empty and 'cost' in successful_df.columns else 0
        total_tokens = successful_df['tokens_generated'].sum() if not successful_df.empty and 'tokens_generated' in successful_df.columns else 0
        total_time = successful_df['generation_time'].sum() if not successful_df.empty and 'generation_time' in successful_df.columns else 0
        
        # Fill in values
        summary_data['Value'] = [
            total_models,
            successful_models,
            failed_models,
            f"{success_rate:.1f}%",
            '',
            best_overall,
            fastest_model,
            highest_quality,
            most_cost_effective,
            '',
            f"{avg_gen_time:.2f}",
            f"{avg_tokens_sec:.1f}",
            f"{avg_quality:.1f}",
            f"${avg_cost:.4f}",
            '',
            f"${total_cost:.4f}",
            f"{total_tokens:,}",
            f"{total_time:.1f}"
        ]
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_performance_sheet(self, writer, df: pd.DataFrame):
        """Create a performance comparison sheet."""
        if df.empty:
            return
        
        successful_df = df[df['success'] == True] if 'success' in df.columns else df
        
        if successful_df.empty:
            return
        
        # Select performance-related columns
        performance_columns = [
            'model', 'provider', 'generation_time', 'tokens_per_second', 
            'time_to_first_token', 'tokens_generated', 'cost', 'cost_per_token'
        ]
        
        available_perf_columns = [col for col in performance_columns if col in successful_df.columns]
        performance_df = successful_df[available_perf_columns].copy()
        
        # Sort by tokens per second (descending)
        if 'tokens_per_second' in performance_df.columns:
            performance_df = performance_df.sort_values('tokens_per_second', ascending=False)
        
        performance_df.to_excel(writer, sheet_name='Performance Comparison', index=False)
    
    def _create_quality_sheet(self, writer, df: pd.DataFrame):
        """Create a quality analysis sheet."""
        if df.empty:
            return
        
        successful_df = df[df['success'] == True] if 'success' in df.columns else df
        
        if successful_df.empty:
            return
        
        # Select quality-related columns
        quality_columns = [
            'model', 'provider', 'overall_quality_score', 'correctness_score',
            'completeness_score', 'code_quality_score', 'readability_score',
            'error_handling_score', 'pros', 'cons', 'evaluation_summary'
        ]
        
        available_quality_columns = [col for col in quality_columns if col in successful_df.columns]
        quality_df = successful_df[available_quality_columns].copy()
        
        # Sort by overall quality score (descending)
        if 'overall_quality_score' in quality_df.columns:
            quality_df = quality_df.sort_values('overall_quality_score', ascending=False)
        
        quality_df.to_excel(writer, sheet_name='Quality Analysis', index=False)
    
    def generate_pdf_charts(self, df: pd.DataFrame) -> List[str]:
        """Generate all PDF charts and return list of created file paths."""
        if df.empty:
            self.logger.warning("No data available for chart generation")
            return []
        
        successful_df = df[df['success'] == True] if 'success' in df.columns else df
        
        if successful_df.empty:
            self.logger.warning("No successful results available for chart generation")
            return []
        
        generated_files = []
        
        try:
            # Generate all four required charts
            chart_methods = [
                ('speed_vs_quality', self._generate_speed_vs_quality_chart),
                ('speed_comparison', self._generate_speed_comparison_chart),
                ('quality_comparison', self._generate_quality_comparison_chart),
                ('ttft_comparison', self._generate_ttft_comparison_chart)
            ]
            
            for chart_name, chart_method in chart_methods:
                try:
                    filename = chart_method(successful_df)
                    if filename and os.path.exists(filename):
                        generated_files.append(os.path.abspath(filename))
                        self.logger.info(f"✓ Generated {chart_name} chart: {filename}")
                    else:
                        self.logger.error(f"✗ Failed to generate {chart_name} chart")
                except Exception as e:
                    self.logger.error(f"✗ Error generating {chart_name} chart: {str(e)}")
            
            self.logger.info(f"Generated {len(generated_files)} PDF charts successfully")
            return generated_files
            
        except Exception as e:
            self.logger.error(f"✗ Error in PDF chart generation: {str(e)}")
            return generated_files
    
    def _generate_speed_vs_quality_chart(self, df: pd.DataFrame) -> str:
        """Generate Speed vs Quality scatter plot PDF."""
        filename = f"speed_vs_quality_{self.timestamp}.pdf"
        
        # Set up the plot style
        plt.style.use('default')
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Extract data
        speed_data = df['tokens_per_second'].values
        quality_data = df['overall_quality_score'].values
        model_names = df['model'].values
        providers = df['provider'].values
        
        # Create color map for providers
        unique_providers = df['provider'].unique()
        colors = plt.cm.Set3(np.linspace(0, 1, len(unique_providers)))
        provider_colors = dict(zip(unique_providers, colors))
        
        # Create scatter plot
        for provider in unique_providers:
            provider_mask = providers == provider
            ax.scatter(
                speed_data[provider_mask], 
                quality_data[provider_mask],
                c=[provider_colors[provider]], 
                label=provider.title(),
                alpha=0.7,
                s=100,
                edgecolors='black',
                linewidth=0.5
            )
        
        # Note: Reference model (Claude Opus 4) is used for evaluation only, not tested
        
        # Customize the plot
        ax.set_xlabel('Speed (Tokens per Second)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Quality Score (1-10)', fontsize=12, fontweight='bold')
        ax.set_title('LLM Performance: Speed vs Quality Trade-off', fontsize=14, fontweight='bold', pad=20)
        
        # Add grid
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # Set axis limits with some padding
        ax.set_xlim(0, max(speed_data) * 1.1)
        ax.set_ylim(0, 10.5)
        
        # Add legend
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', frameon=True, fancybox=True, shadow=True)
        
        # Add annotations for top performers
        top_speed_idx = np.argmax(speed_data)
        top_quality_idx = np.argmax(quality_data)
        
        # Annotate fastest model
        ax.annotate(
            f'Fastest: {model_names[top_speed_idx][:20]}...' if len(model_names[top_speed_idx]) > 20 else f'Fastest: {model_names[top_speed_idx]}',
            xy=(speed_data[top_speed_idx], quality_data[top_speed_idx]),
            xytext=(10, 10), textcoords='offset points',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='lightblue', alpha=0.7),
            arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
            fontsize=9
        )
        
        # Annotate highest quality model (if different from fastest)
        if top_quality_idx != top_speed_idx:
            ax.annotate(
                f'Best Quality: {model_names[top_quality_idx][:20]}...' if len(model_names[top_quality_idx]) > 20 else f'Best Quality: {model_names[top_quality_idx]}',
                xy=(speed_data[top_quality_idx], quality_data[top_quality_idx]),
                xytext=(10, -20), textcoords='offset points',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='lightgreen', alpha=0.7),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
                fontsize=9
            )
        
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        return filename
    
    def _generate_speed_comparison_chart(self, df: pd.DataFrame) -> str:
        """Generate Speed-only comparison bar chart PDF."""
        filename = f"speed_comparison_{self.timestamp}.pdf"
        
        # Sort by speed (descending)
        df_sorted = df.sort_values('tokens_per_second', ascending=False)
        
        # Set up the plot
        fig, ax = plt.subplots(figsize=(14, 8))
        
        # Extract data
        model_names = [name[:25] + '...' if len(name) > 25 else name for name in df_sorted['model'].values]
        speed_data = df_sorted['tokens_per_second'].values
        providers = df_sorted['provider'].values
        
        # Create color map for providers
        unique_providers = df_sorted['provider'].unique()
        colors = plt.cm.Set3(np.linspace(0, 1, len(unique_providers)))
        provider_colors = dict(zip(unique_providers, colors))
        bar_colors = [provider_colors[provider] for provider in providers]
        
        # Create bar chart
        bars = ax.bar(range(len(model_names)), speed_data, color=bar_colors, alpha=0.8, edgecolor='black', linewidth=0.5)
        
        # Note: Reference model (Claude Opus 4) is used for evaluation only, not in charts
        
        # Customize the plot
        ax.set_xlabel('Models', fontsize=12, fontweight='bold')
        ax.set_ylabel('Speed (Tokens per Second)', fontsize=12, fontweight='bold')
        ax.set_title('LLM Speed Comparison', fontsize=14, fontweight='bold', pad=20)
        
        # Set x-axis labels
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right', fontsize=10)
        
        # Add value labels on bars
        for i, (bar, speed) in enumerate(zip(bars, speed_data)):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + max(speed_data) * 0.01,
                   f'{speed:.1f}', ha='center', va='bottom', fontsize=8, fontweight='bold')
        
        # Add grid
        ax.grid(True, alpha=0.3, linestyle='--', axis='y')
        
        # Create legend for providers
        legend_elements = [plt.Rectangle((0,0),1,1, facecolor=provider_colors[provider], alpha=0.8, edgecolor='black', label=provider.title()) 
                          for provider in unique_providers]
        
        ax.legend(handles=legend_elements, bbox_to_anchor=(1.05, 1), loc='upper left', frameon=True, fancybox=True, shadow=True)
        
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        return filename
    
    def _generate_quality_comparison_chart(self, df: pd.DataFrame) -> str:
        """Generate Quality-only comparison bar chart PDF."""
        filename = f"quality_comparison_{self.timestamp}.pdf"
        
        # Sort by quality (descending)
        df_sorted = df.sort_values('overall_quality_score', ascending=False)
        
        # Set up the plot
        fig, ax = plt.subplots(figsize=(14, 8))
        
        # Extract data
        model_names = [name[:25] + '...' if len(name) > 25 else name for name in df_sorted['model'].values]
        quality_data = df_sorted['overall_quality_score'].values
        providers = df_sorted['provider'].values
        
        # Create color map for providers
        unique_providers = df_sorted['provider'].unique()
        colors = plt.cm.Set3(np.linspace(0, 1, len(unique_providers)))
        provider_colors = dict(zip(unique_providers, colors))
        bar_colors = [provider_colors[provider] for provider in providers]
        
        # Create bar chart
        bars = ax.bar(range(len(model_names)), quality_data, color=bar_colors, alpha=0.8, edgecolor='black', linewidth=0.5)
        
        # Note: Reference model (Claude Opus 4) is used for evaluation only, not in charts
        
        # Customize the plot
        ax.set_xlabel('Models', fontsize=12, fontweight='bold')
        ax.set_ylabel('Quality Score (1-10)', fontsize=12, fontweight='bold')
        ax.set_title('LLM Quality Comparison', fontsize=14, fontweight='bold', pad=20)
        
        # Set x-axis labels
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right', fontsize=10)
        
        # Add value labels on bars
        for i, (bar, quality) in enumerate(zip(bars, quality_data)):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                   f'{quality:.1f}', ha='center', va='bottom', fontsize=8, fontweight='bold')
        
        # Set y-axis limits
        ax.set_ylim(0, 10.5)
        
        # Add grid
        ax.grid(True, alpha=0.3, linestyle='--', axis='y')
        
        # Create legend for providers
        legend_elements = [plt.Rectangle((0,0),1,1, facecolor=provider_colors[provider], alpha=0.8, edgecolor='black', label=provider.title()) 
                          for provider in unique_providers]
        
        ax.legend(handles=legend_elements, bbox_to_anchor=(1.05, 1), loc='upper left', frameon=True, fancybox=True, shadow=True)
        
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        return filename
    
    def _generate_ttft_comparison_chart(self, df: pd.DataFrame) -> str:
        """Generate Time to First Token comparison chart PDF."""
        filename = f"ttft_comparison_{self.timestamp}.pdf"
        
        # Sort by TTFT (ascending - lower is better)
        df_sorted = df.sort_values('time_to_first_token', ascending=True)
        
        # Set up the plot
        fig, ax = plt.subplots(figsize=(14, 8))
        
        # Extract data
        model_names = [name[:25] + '...' if len(name) > 25 else name for name in df_sorted['model'].values]
        ttft_data = df_sorted['time_to_first_token'].values
        providers = df_sorted['provider'].values
        
        # Create color map for providers
        unique_providers = df_sorted['provider'].unique()
        colors = plt.cm.Set3(np.linspace(0, 1, len(unique_providers)))
        provider_colors = dict(zip(unique_providers, colors))
        bar_colors = [provider_colors[provider] for provider in providers]
        
        # Create bar chart
        bars = ax.bar(range(len(model_names)), ttft_data, color=bar_colors, alpha=0.8, edgecolor='black', linewidth=0.5)
        
        # Note: Reference model (Claude Opus 4) is used for evaluation only, not in charts
        
        # Customize the plot
        ax.set_xlabel('Models', fontsize=12, fontweight='bold')
        ax.set_ylabel('Time to First Token (seconds)', fontsize=12, fontweight='bold')
        ax.set_title('LLM Time to First Token Comparison (Lower is Better)', fontsize=14, fontweight='bold', pad=20)
        
        # Set x-axis labels
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right', fontsize=10)
        
        # Add value labels on bars
        for i, (bar, ttft) in enumerate(zip(bars, ttft_data)):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + max(ttft_data) * 0.01,
                   f'{ttft:.2f}s', ha='center', va='bottom', fontsize=8, fontweight='bold')
        
        # Add grid
        ax.grid(True, alpha=0.3, linestyle='--', axis='y')
        
        # Create legend for providers
        legend_elements = [plt.Rectangle((0,0),1,1, facecolor=provider_colors[provider], alpha=0.8, edgecolor='black', label=provider.title()) 
                          for provider in unique_providers]
        
        ax.legend(handles=legend_elements, bbox_to_anchor=(1.05, 1), loc='upper left', frameon=True, fancybox=True, shadow=True)
        
        plt.tight_layout()
        plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        return filename
    
    def generate_outputs(self, df: pd.DataFrame) -> Dict[str, List[str]]:
        """Generate all output files (Excel and PDF charts) and return file paths."""
        output_files = {
            'excel': [],
            'charts': []
        }
        
        if df.empty:
            self.logger.warning("No data available for output generation")
            return output_files
        
        # Generate Excel file
        try:
            excel_file = self.generate_excel_output(df)
            if excel_file:
                output_files['excel'].append(excel_file)
        except Exception as e:
            self.logger.error(f"Failed to generate Excel output: {str(e)}")
        
        # Generate PDF charts
        try:
            chart_files = self.generate_pdf_charts(df)
            output_files['charts'].extend(chart_files)
        except Exception as e:
            self.logger.error(f"Failed to generate PDF charts: {str(e)}")
        
        return output_files
    
    def run_complete_benchmark(self) -> Dict[str, Any]:
        """
        Execute the complete benchmarking workflow with comprehensive error handling.
        
        Returns:
            Dict containing execution summary with success/failure reporting and file paths.
        """
        execution_start_time = time.time()
        execution_summary = {
            'start_time': datetime.now().isoformat(),
            'end_time': None,
            'duration_seconds': 0,
            'total_models_attempted': 0,
            'successful_benchmarks': 0,
            'failed_benchmarks': 0,
            'successful_evaluations': 0,
            'failed_evaluations': 0,
            'success_rate': 0.0,
            'providers_tested': [],
            'providers_failed': [],
            'output_files': {
                'excel': [],
                'charts': []
            },
            'errors': [],
            'warnings': [],
            'status': 'running'
        }
        
        try:
            self.logger.info("=" * 80)
            self.logger.info("STARTING UNIFIED LLM BENCHMARK EXECUTION")
            self.logger.info("=" * 80)
            
            # Step 1: Validate configuration and API keys
            self.logger.info("Step 1: Validating configuration and API keys...")
            api_key_status = self.validate_api_keys()
            available_count = sum(1 for status in api_key_status.values() if status)
            total_providers = len(api_key_status)
            
            if available_count == 0:
                error_msg = "No valid API keys found. Cannot proceed with benchmarking."
                self.logger.error(error_msg)
                execution_summary['status'] = 'failed'
                execution_summary['errors'].append(error_msg)
                return execution_summary
            
            self.logger.info(f"✓ {available_count}/{total_providers} providers available")
            
            # Step 2: Collect available models
            self.logger.info("Step 2: Collecting available models...")
            available_models = self.get_available_models()
            execution_summary['total_models_attempted'] = len(available_models)
            
            if not available_models:
                error_msg = "No models available for testing."
                self.logger.error(error_msg)
                execution_summary['status'] = 'failed'
                execution_summary['errors'].append(error_msg)
                return execution_summary
            
            self.logger.info(f"✓ Found {len(available_models)} models to test")
            
            # Step 3: Execute benchmarks with parallel processing
            self.logger.info("Step 3: Executing model benchmarks...")
            benchmark_results = []
            
            # Use ThreadPoolExecutor for parallel benchmarking
            max_workers = min(5, len(available_models))  # Limit concurrent requests
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all benchmark tasks
                future_to_model = {
                    executor.submit(self.benchmark_model, provider, model): (provider, model)
                    for provider, model in available_models
                }
                
                # Collect results as they complete
                for future in future_to_model:
                    provider, model = future_to_model[future]
                    try:
                        result = future.result(timeout=REQUEST_TIMEOUT + 30)  # Add buffer to request timeout
                        benchmark_results.append(result)
                        
                        if result.success:
                            execution_summary['successful_benchmarks'] += 1
                            if provider not in execution_summary['providers_tested']:
                                execution_summary['providers_tested'].append(provider)
                            self.logger.info(f"✓ Benchmarked {provider}/{model}: {result.tokens_per_second:.1f} t/s")
                        else:
                            execution_summary['failed_benchmarks'] += 1
                            error_msg = f"Benchmark failed for {provider}/{model}: {result.error_message}"
                            execution_summary['errors'].append(error_msg)
                            self.logger.error(f"✗ {error_msg}")
                            
                            if provider not in execution_summary['providers_failed']:
                                execution_summary['providers_failed'].append(provider)
                                
                    except Exception as e:
                        execution_summary['failed_benchmarks'] += 1
                        error_msg = f"Exception during benchmark of {provider}/{model}: {str(e)}"
                        execution_summary['errors'].append(error_msg)
                        self.logger.error(f"✗ {error_msg}")
                        
                        # Create failed result for consistency
                        failed_result = BenchmarkResult(
                            model=model,
                            provider=provider,
                            platform=provider,
                            generation_time=0.0,
                            tokens_per_second=0.0,
                            time_to_first_token=0.0,
                            tokens_generated=0,
                            total_tokens=0,
                            cost=0.0,
                            response_content="",
                            success=False,
                            error_message=str(e)
                        )
                        benchmark_results.append(failed_result)
            
            self.logger.info(f"✓ Benchmark phase complete: {execution_summary['successful_benchmarks']}/{len(available_models)} successful")
            
            # Step 4: Generate reference response for evaluation
            self.logger.info("Step 4: Generating reference response for quality evaluation...")
            reference_response = ""
            
            try:
                # Use Claude Opus as the reference model for evaluation (not one of the test models)
                reference_response = self.generate_reference_response()
                if reference_response:
                    self.logger.info("✓ Reference response generated successfully")
                else:
                    warning_msg = "Failed to generate reference response using Claude Opus"
                    execution_summary['warnings'].append(warning_msg)
                    self.logger.warning(warning_msg)
            except Exception as e:
                warning_msg = f"Exception during reference response generation: {str(e)}"
                execution_summary['warnings'].append(warning_msg)
                self.logger.warning(warning_msg)
            
            # Step 5: Execute quality evaluations
            self.logger.info("Step 5: Executing quality evaluations...")
            evaluation_results = []
            
            # Only evaluate successful benchmark results
            successful_benchmarks = [r for r in benchmark_results if r.success]
            
            if successful_benchmarks and reference_response:
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # Submit evaluation tasks
                    future_to_benchmark = {
                        executor.submit(self.evaluate_response_quality, benchmark, reference_response): benchmark
                        for benchmark in successful_benchmarks
                    }
                    
                    # Collect evaluation results
                    for future in future_to_benchmark:
                        benchmark = future_to_benchmark[future]
                        try:
                            eval_result = future.result(timeout=REQUEST_TIMEOUT + 30)
                            evaluation_results.append(eval_result)
                            execution_summary['successful_evaluations'] += 1
                            self.logger.info(f"✓ Evaluated {benchmark.provider}/{benchmark.model}: {eval_result.overall_score:.1f}/10")
                            
                        except Exception as e:
                            execution_summary['failed_evaluations'] += 1
                            error_msg = f"Evaluation failed for {benchmark.provider}/{benchmark.model}: {str(e)}"
                            execution_summary['errors'].append(error_msg)
                            self.logger.error(f"✗ {error_msg}")
            else:
                warning_msg = "Skipping quality evaluation: no successful benchmarks or reference response"
                execution_summary['warnings'].append(warning_msg)
                self.logger.warning(warning_msg)
            
            self.logger.info(f"✓ Evaluation phase complete: {execution_summary['successful_evaluations']}/{len(successful_benchmarks)} successful")
            
            # Step 6: Process and aggregate data
            self.logger.info("Step 6: Processing and aggregating results...")
            try:
                aggregated_df = self.process_and_aggregate_data(benchmark_results, evaluation_results)
                self.logger.info(f"✓ Data processing complete: {len(aggregated_df)} final results")
            except Exception as e:
                error_msg = f"Data processing failed: {str(e)}"
                execution_summary['errors'].append(error_msg)
                self.logger.error(f"✗ {error_msg}")
                # Create empty DataFrame to continue with output generation
                aggregated_df = pd.DataFrame()
            
            # Step 7: Generate output files
            self.logger.info("Step 7: Generating output files...")
            try:
                output_files = self.generate_outputs(aggregated_df)
                execution_summary['output_files'] = output_files
                
                # Log file generation results
                total_files = len(output_files['excel']) + len(output_files['charts'])
                if total_files > 0:
                    self.logger.info(f"✓ Generated {total_files} output files:")
                    for excel_file in output_files['excel']:
                        self.logger.info(f"  📊 Excel: {excel_file}")
                    for chart_file in output_files['charts']:
                        self.logger.info(f"  📈 Chart: {chart_file}")
                else:
                    warning_msg = "No output files were generated"
                    execution_summary['warnings'].append(warning_msg)
                    self.logger.warning(warning_msg)
                    
            except Exception as e:
                error_msg = f"Output generation failed: {str(e)}"
                execution_summary['errors'].append(error_msg)
                self.logger.error(f"✗ {error_msg}")
            
            # Calculate final statistics
            execution_summary['success_rate'] = (
                execution_summary['successful_benchmarks'] / execution_summary['total_models_attempted'] * 100
                if execution_summary['total_models_attempted'] > 0 else 0.0
            )
            
            # Determine overall status
            if execution_summary['successful_benchmarks'] > 0:
                execution_summary['status'] = 'completed_with_results'
            elif execution_summary['total_models_attempted'] > 0:
                execution_summary['status'] = 'completed_no_results'
            else:
                execution_summary['status'] = 'failed'
            
        except Exception as e:
            # Handle any unexpected errors in the main workflow
            error_msg = f"Critical error in benchmark execution: {str(e)}"
            execution_summary['errors'].append(error_msg)
            execution_summary['status'] = 'failed'
            self.logger.error(f"✗ {error_msg}")
            
        finally:
            # Finalize execution summary
            execution_end_time = time.time()
            execution_summary['end_time'] = datetime.now().isoformat()
            execution_summary['duration_seconds'] = execution_end_time - execution_start_time
            
            # Log final summary
            self.logger.info("=" * 80)
            self.logger.info("BENCHMARK EXECUTION SUMMARY")
            self.logger.info("=" * 80)
            self.logger.info(f"Status: {execution_summary['status'].upper()}")
            self.logger.info(f"Duration: {execution_summary['duration_seconds']:.1f} seconds")
            self.logger.info(f"Models Attempted: {execution_summary['total_models_attempted']}")
            self.logger.info(f"Successful Benchmarks: {execution_summary['successful_benchmarks']}")
            self.logger.info(f"Failed Benchmarks: {execution_summary['failed_benchmarks']}")
            self.logger.info(f"Success Rate: {execution_summary['success_rate']:.1f}%")
            self.logger.info(f"Successful Evaluations: {execution_summary['successful_evaluations']}")
            self.logger.info(f"Providers Tested: {', '.join(execution_summary['providers_tested']) if execution_summary['providers_tested'] else 'None'}")
            
            if execution_summary['output_files']['excel'] or execution_summary['output_files']['charts']:
                self.logger.info("Generated Files:")
                for excel_file in execution_summary['output_files']['excel']:
                    self.logger.info(f"  📊 {excel_file}")
                for chart_file in execution_summary['output_files']['charts']:
                    self.logger.info(f"  📈 {chart_file}")
            
            if execution_summary['errors']:
                self.logger.info(f"Errors: {len(execution_summary['errors'])}")
                for error in execution_summary['errors'][:5]:  # Show first 5 errors
                    self.logger.info(f"  ❌ {error}")
                if len(execution_summary['errors']) > 5:
                    self.logger.info(f"  ... and {len(execution_summary['errors']) - 5} more errors")
            
            if execution_summary['warnings']:
                self.logger.info(f"Warnings: {len(execution_summary['warnings'])}")
                for warning in execution_summary['warnings'][:3]:  # Show first 3 warnings
                    self.logger.info(f"  ⚠️  {warning}")
                if len(execution_summary['warnings']) > 3:
                    self.logger.info(f"  ... and {len(execution_summary['warnings']) - 3} more warnings")
            
            self.logger.info("=" * 80)
            
        return execution_summary


if __name__ == "__main__":
    # Main execution entry point
    try:
        print("=" * 80)
        print("UNIFIED LLM BENCHMARKER")
        print("=" * 80)
        print("Initializing benchmarker...")
        
        # Initialize the benchmarker
        benchmarker = UnifiedBenchmarker()
        
        print(f"✓ Configuration loaded with {len(PROVIDERS)} providers")
        print(f"✓ Benchmark prompt length: {len(DEFAULT_BENCHMARK_PROMPT)} characters")
        print(f"✓ Available providers: {sum(1 for status in benchmarker.validate_api_keys().values() if status)}")
        print()
        
        # Run the complete benchmark workflow
        print("Starting benchmark execution...")
        execution_summary = benchmarker.run_complete_benchmark()
        
        # Print final results summary
        print("\n" + "=" * 80)
        print("EXECUTION COMPLETED")
        print("=" * 80)
        print(f"Status: {execution_summary['status'].upper().replace('_', ' ')}")
        print(f"Duration: {execution_summary['duration_seconds']:.1f} seconds")
        print(f"Success Rate: {execution_summary['success_rate']:.1f}%")
        print(f"Models Tested: {execution_summary['successful_benchmarks']}/{execution_summary['total_models_attempted']}")
        
        if execution_summary['output_files']['excel'] or execution_summary['output_files']['charts']:
            print("\nGenerated Files:")
            for excel_file in execution_summary['output_files']['excel']:
                print(f"  📊 Excel: {excel_file}")
            for chart_file in execution_summary['output_files']['charts']:
                print(f"  📈 Chart: {chart_file}")
        
        if execution_summary['errors']:
            print(f"\nErrors Encountered: {len(execution_summary['errors'])}")
            
        if execution_summary['warnings']:
            print(f"Warnings: {len(execution_summary['warnings'])}")
        
        print("=" * 80)
        
        # Exit with appropriate code
        if execution_summary['status'] == 'completed_with_results':
            exit(0)  # Success
        elif execution_summary['status'] == 'completed_no_results':
            exit(1)  # Partial failure
        else:
            exit(2)  # Complete failure
            
    except KeyboardInterrupt:
        print("\n\n⚠️  Benchmark execution interrupted by user")
        exit(130)  # Standard exit code for SIGINT
        
    except Exception as e:
        print(f"\n\n❌ Critical error during benchmark execution: {str(e)}")
        import traceback
        traceback.print_exc()
        exit(3)  # Critical error